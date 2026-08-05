"""jp_tostnet.py — JPX ToSTNeT 超大口約定情報（≥50 億円 block trade）scraper.

JPX 免费公开数据，T+1 发布。用于印证 moomoo capital flow anomaly 或
夜盘/私聊传来的 "大买家" 消息是否真的落地为 block trade。

用法：
  python jp_tostnet.py                # 拉最近 5 日报告 → 保存到 signals/jp_tostnet/
  python jp_tostnet.py --days 10      # 拉最近 10 日
  python jp_tostnet.py --verify 6981  # 查最近 10 日里 6981 是否出现

数据源：
  index: https://www.jpx.co.jp/markets/equities/tostnet/index.html
  daily: /markets/equities/tostnet/<random-id>-att/YYYYMMDD_ToSTNeT_Trading_Information.xlsx
"""
from __future__ import annotations

import io
import json
import re
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Optional

BASE_DIR = Path(__file__).parent
CACHE_DIR = BASE_DIR / "signals" / "jp_tostnet"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

JPX_INDEX_URL = "https://www.jpx.co.jp/markets/equities/tostnet/index.html"
JPX_BASE = "https://www.jpx.co.jp"

_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0 Safari/537.36"
_HEADERS = {"User-Agent": _UA, "Accept-Language": "ja,en;q=0.9"}


def _http_get(url: str, timeout: int = 20) -> bytes:
    req = urllib.request.Request(url, headers=_HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def fetch_index_links() -> list[dict]:
    """Parse index page → [{date_str: '20260803', url: '/markets/.../20260803_...xlsx'}, ...]"""
    try:
        html = _http_get(JPX_INDEX_URL).decode("utf-8", errors="replace")
    except Exception as e:
        return [{"error": f"index_fetch: {e}"}]
    pattern = r'href=["\']([^"\']+/(\d{8})_ToSTNeT_Trading_Information\.xlsx)["\']'
    out = []
    for m in re.finditer(pattern, html):
        rel_url, date_str = m.group(1), m.group(2)
        url = rel_url if rel_url.startswith("http") else f"{JPX_BASE}{rel_url}"
        out.append({"date": date_str, "url": url})
    # dedup by date, keep first
    seen: set[str] = set()
    unique = []
    for e in out:
        if e["date"] not in seen:
            seen.add(e["date"])
            unique.append(e)
    return unique


def _parse_xlsx(data: bytes) -> list[dict]:
    """Parse ToSTNeT daily xlsx → list of trade rows."""
    try:
        import openpyxl
    except Exception:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    trades = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Row 0 = headers (mixed ja/en)
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            try:
                code = str(row[3]).strip() if row[3] is not None else ""
                if not code:
                    continue
                price = float(str(row[6]).replace(",", "")) if row[6] else 0
                volume = int(str(row[7]).replace(",", "")) if row[7] else 0
                value_yen = int(str(row[8]).replace(",", "")) if row[8] else 0
                trades.append({
                    "publication_date": str(row[0]),
                    "trading_date":     str(row[1]),
                    "trade_time":       str(row[2]),
                    "code":             code,
                    "name_ja":          str(row[4]) if row[4] else "",
                    "name_en":          str(row[5]) if row[5] else "",
                    "price_yen":        price,
                    "volume_shares":    volume,
                    "value_yen":        value_yen,
                    "value_oku":        round(value_yen / 1e8, 1),   # 億円 unit
                    "sheet":            sheet_name,
                })
            except (ValueError, TypeError, IndexError):
                continue
    return trades


def fetch_daily_report(date_yyyymmdd: str, force: bool = False) -> Optional[dict]:
    """拉某日 ToSTNeT 报告 → {trades: [...], count, total_value_oku}. Cached to disk."""
    cache_path = CACHE_DIR / f"{date_yyyymmdd}.json"
    if cache_path.exists() and not force:
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    idx = fetch_index_links()
    if not idx or (idx and idx[0].get("error")):
        return {"error": idx[0].get("error") if idx else "no_index", "date": date_yyyymmdd}
    entry = next((e for e in idx if e["date"] == date_yyyymmdd), None)
    if not entry:
        return {"error": "date_not_available", "date": date_yyyymmdd,
                "available_dates": [e["date"] for e in idx[:10]]}
    try:
        data = _http_get(entry["url"], timeout=30)
    except Exception as e:
        return {"error": f"download: {e}", "date": date_yyyymmdd}
    trades = _parse_xlsx(data)
    result = {
        "date":            date_yyyymmdd,
        "source_url":      entry["url"],
        "fetched_at":      datetime.now().isoformat(timespec="seconds"),
        "count":           len(trades),
        "total_value_oku": round(sum(t["value_yen"] for t in trades) / 1e8, 1),
        "trades":          trades,
    }
    try:
        cache_path.write_text(json.dumps(result, ensure_ascii=False, indent=2),
                              encoding="utf-8")
    except Exception:
        pass
    return result


def fetch_recent(n_days: int = 5) -> list[dict]:
    """拉最近 n_days 日的 ToSTNeT 报告。返回 list of daily result dicts."""
    idx = fetch_index_links()
    if not idx or (idx and idx[0].get("error")):
        return [{"error": idx[0].get("error") if idx else "no_index"}]
    out = []
    for entry in idx[:n_days]:
        out.append(fetch_daily_report(entry["date"]))
    return out


def check_ticker_hits(ticker: str, n_days: int = 10) -> list[dict]:
    """查最近 n_days 日 ToSTNeT 里某个 ticker 是否出现。

    ticker: JP code (4-digit like '6981' or '6981.T' or 'JP.6981')
    """
    code = re.sub(r"\.T$|^JP\.", "", ticker.upper().strip())
    reports = fetch_recent(n_days)
    hits = []
    for rep in reports:
        if not rep or rep.get("error"):
            continue
        for t in rep.get("trades", []):
            if t.get("code") == code:
                hits.append({
                    "date":        rep["date"],
                    "trade_time":  t["trade_time"],
                    "price":       t["price_yen"],
                    "volume":      t["volume_shares"],
                    "value_oku":   t["value_oku"],
                    "sheet":       t["sheet"],
                })
    return hits


def summary_for_watchlist(watchlist_codes: list[str], n_days: int = 5) -> dict:
    """给 JP watchlist 生成印证摘要：每个 ticker 最近 n 日是否命中 ToSTNeT。"""
    reports = fetch_recent(n_days)
    hits_by_code: dict[str, list] = {code: [] for code in watchlist_codes}
    for rep in reports:
        if not rep or rep.get("error"):
            continue
        for t in rep.get("trades", []):
            c = t.get("code")
            if c in hits_by_code:
                hits_by_code[c].append({
                    "date":       rep["date"],
                    "trade_time": t["trade_time"],
                    "price":      t["price_yen"],
                    "volume":     t["volume_shares"],
                    "value_oku":  t["value_oku"],
                })
    return {
        "checked_at":       datetime.now().isoformat(timespec="seconds"),
        "days_covered":     [r["date"] for r in reports if r and not r.get("error")],
        "hits_by_code":     hits_by_code,
        "any_hit":          any(v for v in hits_by_code.values()),
    }


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--days", type=int, default=5)
    p.add_argument("--verify", type=str, default=None,
                   help="Ticker code (e.g. 6981) 查最近 N 日是否命中 ToSTNeT")
    p.add_argument("--force", action="store_true", help="忽略缓存重下")
    args = p.parse_args()

    if args.verify:
        hits = check_ticker_hits(args.verify, n_days=args.days)
        print(json.dumps(hits, ensure_ascii=False, indent=2))
    else:
        for rep in fetch_recent(args.days):
            if rep.get("error"):
                print(f"⚠ {rep}")
                continue
            print(f"{rep['date']}: {rep['count']} trades, {rep['total_value_oku']}億円 total")
            for t in rep.get("trades", [])[:5]:
                print(f"  · {t['trade_time']} [{t['code']}] {t['name_ja']} "
                      f"{t['volume_shares']:,}@¥{t['price_yen']:,} = {t['value_oku']}億円")
